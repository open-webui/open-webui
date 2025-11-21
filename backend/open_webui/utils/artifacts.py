"""
Utilities for emitting various artifact types to the frontend via event emitters.
"""

import logging
import uuid
from pathlib import Path
from typing import Optional, Dict, Any, List
from open_webui.models.files import FileModel, FileForm, Files

log = logging.getLogger(__name__)


async def emit_excel_artifact(
    event_emitter,
    file_model: FileModel,
    webui_url: str,
    sheet_names: Optional[List[str]] = None,
    active_sheet: Optional[str] = None,
):
    """
    Emit a 'files' event with a single Excel artifact for the given file_model.

    Args:
        event_emitter: The event emitter function to send events to the frontend
        file_model: The FileModel instance representing the Excel file
        webui_url: Base URL of the WebUI instance (e.g., "http://localhost:3000")
        sheet_names: Optional list of sheet names in the workbook
        active_sheet: Optional name of the active/default sheet to display
    """
    try:
        # Construct the download URL for the file
        download_url = f"{webui_url}/api/v1/files/{file_model.id}/content"

        # Build metadata
        meta: Dict[str, Any] = {}
        if sheet_names:
            meta["sheetNames"] = sheet_names
        if active_sheet:
            meta["activeSheet"] = active_sheet
        elif sheet_names and len(sheet_names) > 0:
            # Default to first sheet if not specified
            meta["activeSheet"] = sheet_names[0]

        # Emit the files event
        await event_emitter(
            {
                "type": "files",
                "data": {
                    "files": [
                        {
                            "type": "excel",
                            "url": download_url,
                            "name": file_model.filename,
                            "fileId": str(file_model.id),
                            "meta": meta if meta else {},
                        }
                    ]
                },
            }
        )

        log.info(f"Excel artifact emitted for file {file_model.id}: {file_model.filename}")

    except Exception as e:
        log.error(f"Error emitting Excel artifact: {e}")
        raise


async def emit_file_artifacts(
    event_emitter,
    file_models: List[FileModel],
    webui_url: str,
):
    """
    Emit a 'files' event with multiple file artifacts.

    Args:
        event_emitter: The event emitter function to send events to the frontend
        file_models: List of FileModel instances
        webui_url: Base URL of the WebUI instance
    """
    try:
        files = []
        for file_model in file_models:
            download_url = f"{webui_url}/api/v1/files/{file_model.id}/content"

            # Determine file type from content_type or extension
            content_type = file_model.meta.get("content_type", "") if file_model.meta else ""
            file_type = "file"  # default

            if "image" in content_type:
                file_type = "image"
            elif "audio" in content_type:
                file_type = "audio"
            elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                file_type = "excel"

            files.append({
                "type": file_type,
                "url": download_url,
                "name": file_model.filename,
                "fileId": str(file_model.id),
                "meta": file_model.meta or {},
            })

        await event_emitter(
            {
                "type": "files",
                "data": {
                    "files": files,
                },
            }
        )

        log.info(f"Emitted {len(files)} file artifacts")

    except Exception as e:
        log.error(f"Error emitting file artifacts: {e}")
        raise


def create_excel_file_record(
    user_id: str,
    file_path: str,
    filename: str,
    sheet_names: Optional[List[str]] = None,
) -> Optional[FileModel]:
    """
    Create a file record for an Excel file in the database.

    Args:
        user_id: ID of the user who owns the file
        file_path: Path to the Excel file in storage
        filename: Name of the Excel file
        sheet_names: Optional list of sheet names in the workbook

    Returns:
        FileModel instance if successful, None otherwise

    Example usage in a tool:
        ```python
        import openpyxl
        from open_webui.utils.artifacts import create_excel_file_record

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws['A1'] = "Hello"
        ws['B1'] = "World"

        # Save to temp file
        temp_path = f"/tmp/report_{uuid.uuid4()}.xlsx"
        wb.save(temp_path)

        # Create file record
        file_record = create_excel_file_record(
            user_id=user.id,
            file_path=temp_path,
            filename="report.xlsx",
            sheet_names=wb.sheetnames
        )

        # Return for tool result
        return {
            "type": "excel",
            "fileId": file_record.id,
            "url": f"/api/v1/files/{file_record.id}/content",
            "name": filename
        }
        ```
    """
    try:
        # Create metadata
        meta = {
            "name": filename,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size": Path(file_path).stat().st_size if Path(file_path).exists() else 0,
        }

        if sheet_names:
            meta["sheetNames"] = sheet_names

        # Create file form
        file_form = FileForm(
            id=str(uuid.uuid4()),
            filename=filename,
            path=file_path,
            meta=meta,
            data={"status": "completed"},
        )

        # Insert into database
        file_model = Files.insert_new_file(user_id, file_form)

        if file_model:
            log.info(f"Created Excel file record: {file_model.id} - {filename}")
            return file_model
        else:
            log.error(f"Failed to create Excel file record for {filename}")
            return None

    except Exception as e:
        log.error(f"Error creating Excel file record: {e}")
        return None
