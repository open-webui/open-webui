import black
import logging
import markdown
import re
import time
from datetime import datetime
from io import BytesIO
import csv
import json

from open_webui.models.chats import ChatTitleMessagesForm
from open_webui.config import DATA_DIR, ENABLE_ADMIN_EXPORT
from open_webui.constants import ERROR_MESSAGES
from fastapi import APIRouter, Depends, HTTPException, Request, Response, status
from pydantic import BaseModel
from starlette.responses import FileResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from open_webui.utils.misc import get_gravatar_url
from open_webui.utils.pdf_generator import PDFGenerator
from open_webui.utils.word_generator import WordGenerator
from open_webui.utils.auth import get_admin_user, get_verified_user
from open_webui.utils.code_interpreter import execute_code_jupyter

log = logging.getLogger(__name__)

router = APIRouter()


@router.get('/gravatar')
async def get_gravatar(email: str, user=Depends(get_verified_user)):
    return get_gravatar_url(email)


class CodeForm(BaseModel):
    code: str


@router.post('/code/format')
async def format_code(form_data: CodeForm, user=Depends(get_admin_user)):
    try:
        formatted_code = black.format_str(form_data.code, mode=black.Mode())
        return {'code': formatted_code}
    except black.NothingChanged:
        return {'code': form_data.code}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post('/code/execute')
async def execute_code(request: Request, form_data: CodeForm, user=Depends(get_verified_user)):
    if not request.app.state.config.ENABLE_CODE_EXECUTION:
        raise HTTPException(
            status_code=403,
            detail=ERROR_MESSAGES.FEATURE_DISABLED('Code execution'),
        )

    if request.app.state.config.CODE_EXECUTION_ENGINE == 'jupyter':
        output = await execute_code_jupyter(
            request.app.state.config.CODE_EXECUTION_JUPYTER_URL,
            form_data.code,
            (
                request.app.state.config.CODE_EXECUTION_JUPYTER_AUTH_TOKEN
                if request.app.state.config.CODE_EXECUTION_JUPYTER_AUTH == 'token'
                else None
            ),
            (
                request.app.state.config.CODE_EXECUTION_JUPYTER_AUTH_PASSWORD
                if request.app.state.config.CODE_EXECUTION_JUPYTER_AUTH == 'password'
                else None
            ),
            request.app.state.config.CODE_EXECUTION_JUPYTER_TIMEOUT,
        )

        return output
    else:
        raise HTTPException(
            status_code=400,
            detail=ERROR_MESSAGES.DEFAULT('Code execution engine not supported'),
        )


class MarkdownForm(BaseModel):
    md: str


@router.post('/markdown')
async def get_html_from_markdown(form_data: MarkdownForm, user=Depends(get_verified_user)):
    return {'html': markdown.markdown(form_data.md)}


class ChatForm(BaseModel):
    title: str
    messages: list[dict]


@router.post('/pdf')
async def download_chat_as_pdf(form_data: ChatTitleMessagesForm, user=Depends(get_verified_user)):
    """
    Generate a professional PDF export of a chat conversation using ReportLab.
    Local: emits Content-Disposition with sanitized title + timestamp,
    plus diagnostic X-PDF-Size / X-Generation-Time headers.
    """
    try:
        start_time = time.time()

        # Generate PDF using modern ReportLab generator
        pdf_bytes = PDFGenerator(form_data).generate_chat_pdf()

        elapsed = time.time() - start_time
        size_kb = len(pdf_bytes) / 1024

        log.info(
            f"PDF generated for user {user.id}: "
            f"{len(form_data.messages)} messages, "
            f"{size_kb:.1f} KB, "
            f"{elapsed:.2f}s"
        )

        # Sanitize filename (remove special characters) and add timestamp
        safe_title = re.sub(r'[^\w\s-]', '', form_data.title)[:50]
        safe_title = safe_title.strip() or "chat"

        # Add timestamp to make each export unique
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_title}_{timestamp}.pdf"

        return Response(
            content=pdf_bytes,
            media_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'X-PDF-Size': str(len(pdf_bytes)),
                'X-Generation-Time': f'{elapsed:.2f}s',
            },
        )
    except Exception as e:
        log.exception(f'Error generating PDF for user {user.id}: {e}')
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail='Failed to generate PDF export',
        )


@router.post("/word")
async def download_chat_as_word(form_data: ChatTitleMessagesForm, user=Depends(get_verified_user)):
    """
    Generate a professional Word document export of a chat conversation.

    Uses python-docx for high-quality output with:
    - Proper markdown rendering (headers, code blocks, lists, tables)
    - Professional styling (color-coded messages)
    - Editable Word documents (users can edit after export)
    - Small file sizes (<500KB for 100 messages)
    - Fast generation (<2s for 100 messages)

    Args:
        form_data: Chat title and messages
        user: Authenticated user

    Returns:
        Word document file download
    """
    try:
        start_time = time.time()

        # Generate Word document using python-docx generator
        docx_bytes = WordGenerator(form_data).generate_chat_docx()

        elapsed = time.time() - start_time
        size_kb = len(docx_bytes) / 1024

        log.info(
            f"Word doc generated for user {user.id}: "
            f"{len(form_data.messages)} messages, "
            f"{size_kb:.1f} KB, "
            f"{elapsed:.2f}s"
        )

        # Sanitize filename (remove special characters) and add timestamp
        safe_title = re.sub(r'[^\w\s-]', '', form_data.title)[:50]
        safe_title = safe_title.strip() or "chat"

        # Add timestamp to make each export unique
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_title}_{timestamp}.docx"

        return Response(
            content=docx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "X-Word-Size": str(len(docx_bytes)),
                "X-Generation-Time": f"{elapsed:.2f}s",
            },
        )
    except Exception as e:
        log.exception(f"Error generating Word doc for user {user.id}: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to generate Word export")


class ArtifactExportForm(BaseModel):
    artifact_type: str  # "csv", "json"
    content: str
    filename: str = "data.xlsx"


@router.post("/artifacts/export-excel")
async def export_artifact_to_excel(form_data: ArtifactExportForm, user=Depends(get_verified_user)):
    """
    Convert artifact data (CSV/JSON) to Excel format.

    Supports:
    - CSV data: Parses and converts to Excel with formatting
    - JSON arrays: Converts array of objects to Excel table

    Args:
        form_data: Artifact type, content, and desired filename
        user: Authenticated user

    Returns:
        Excel file download (.xlsx)
    """
    try:
        start_time = time.time()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2d4a5a", end_color="2d4a5a", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")

        if form_data.artifact_type == "csv":
            # Parse CSV and write to worksheet
            reader = csv.reader(form_data.content.splitlines())
            rows = list(reader)

            if rows:
                # Write data
                for row_idx, row in enumerate(rows, start=1):
                    for col_idx, cell_value in enumerate(row, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                        # Style header row
                        if row_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        elif form_data.artifact_type == "json":
            # Parse JSON array and write to worksheet
            try:
                data = json.loads(form_data.content)

                if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                    # Extract headers from first object
                    headers = list(data[0].keys())

                    # Write headers
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # Write data rows
                    for row_idx, item in enumerate(data, start=2):
                        for col_idx, header in enumerate(headers, start=1):
                            value = item.get(header, "")
                            ws.cell(row=row_idx, column=col_idx, value=value)

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                else:
                    raise ValueError("JSON must be an array of objects")

            except json.JSONDecodeError as e:
                log.error(f"Invalid JSON data: {e}")
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON format")
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail=f"Unsupported artifact type: {form_data.artifact_type}"
            )

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_bytes = buffer.getvalue()
        buffer.close()

        elapsed = time.time() - start_time
        size_kb = len(excel_bytes) / 1024

        log.info(
            f"Excel generated for user {user.id}: "
            f"type={form_data.artifact_type}, "
            f"{size_kb:.1f} KB, "
            f"{elapsed:.2f}s"
        )

        # Sanitize filename
        safe_filename = re.sub(r'[^\w\s.-]', '', form_data.filename)
        if not safe_filename.endswith('.xlsx'):
            safe_filename += '.xlsx'

        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_filename}"',
                "X-Excel-Size": str(len(excel_bytes)),
                "X-Generation-Time": f"{elapsed:.2f}s",
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"Error generating Excel for user {user.id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail='Failed to generate Excel export',
        )


@router.get('/db/download')
async def download_db(user=Depends(get_admin_user)):
    if not ENABLE_ADMIN_EXPORT:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
        )
    from open_webui.internal.db import engine

    if engine.name != 'sqlite':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ERROR_MESSAGES.DB_NOT_SQLITE,
        )
    return FileResponse(
        engine.url.database,
        media_type='application/octet-stream',
        filename='webui.db',
    )
