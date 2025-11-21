"""
Excel file editing API endpoints
"""

import logging
from pathlib import Path
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
import openpyxl
from openpyxl.utils import get_column_letter

from open_webui.constants import ERROR_MESSAGES
from open_webui.models.files import Files, FileModel
from open_webui.storage.provider import Storage
from open_webui.utils.auth import get_verified_user

log = logging.getLogger(__name__)

router = APIRouter()


class CellChange(BaseModel):
    """Represents a single cell change"""
    row: int  # 1-based row number
    col: int  # 1-based column number
    value: Optional[str | int | float | bool] = None


class ExcelUpdateRequest(BaseModel):
    """Request to update cells in an Excel file"""
    fileId: str
    sheet: str
    changes: List[CellChange]


class ExcelUpdateResponse(BaseModel):
    """Response from Excel update operation"""
    status: str
    message: Optional[str] = None


class ExcelMetadataResponse(BaseModel):
    """Response with Excel file metadata"""
    fileId: str
    sheetNames: List[str]
    activeSheet: Optional[str] = None


@router.post("/update")
async def update_excel_file(
    request: ExcelUpdateRequest,
    user=Depends(get_verified_user),
) -> ExcelUpdateResponse:
    """
    Update cells in an Excel workbook.

    Args:
        request: The update request with file ID, sheet name, and cell changes
        user: The authenticated user

    Returns:
        ExcelUpdateResponse with status

    Raises:
        HTTPException: If file not found, access denied, or update fails
    """
    try:
        # Get the file from database
        file = Files.get_file_by_id(request.fileId)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        # Check user permissions
        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        # Get the file path from storage
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        # Load the workbook
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            log.error(f"Error loading workbook {request.fileId}: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {str(e)}",
            )

        # Check if sheet exists
        if request.sheet not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{request.sheet}' not found in workbook. Available sheets: {', '.join(wb.sheetnames)}",
            )

        # Get the worksheet
        ws = wb[request.sheet]

        # Apply changes
        changes_applied = 0
        for change in request.changes:
            try:
                # Validate row and column are positive
                if change.row < 1 or change.col < 1:
                    log.warning(f"Invalid cell coordinates: row={change.row}, col={change.col}")
                    continue

                # Get the cell and set its value
                cell = ws.cell(row=change.row, column=change.col)
                cell.value = change.value
                changes_applied += 1

                log.debug(f"Updated cell {get_column_letter(change.col)}{change.row} = {change.value}")

            except Exception as e:
                log.error(f"Error updating cell at row={change.row}, col={change.col}: {e}")
                continue

        # Save the workbook
        try:
            wb.save(file_path)
            wb.close()
            log.info(f"Successfully updated {changes_applied} cells in file {request.fileId}")
        except Exception as e:
            log.error(f"Error saving workbook {request.fileId}: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to save changes: {str(e)}",
            )

        return ExcelUpdateResponse(
            status="ok",
            message=f"Successfully updated {changes_applied} cells",
        )

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error updating Excel file: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )


@router.get("/{file_id}/metadata")
async def get_excel_metadata(
    file_id: str,
    user=Depends(get_verified_user),
) -> ExcelMetadataResponse:
    """
    Get metadata about an Excel file (sheet names, etc.).

    Args:
        file_id: The ID of the Excel file
        user: The authenticated user

    Returns:
        ExcelMetadataResponse with sheet names and active sheet

    Raises:
        HTTPException: If file not found or access denied
    """
    try:
        # Get the file from database
        file = Files.get_file_by_id(file_id)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        # Check user permissions
        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        # Get the file path from storage
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        # Load the workbook (data_only=True for faster loading)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            active_sheet = wb.active.title if wb.active else None
            wb.close()

            return ExcelMetadataResponse(
                fileId=file_id,
                sheetNames=sheet_names,
                activeSheet=active_sheet,
            )

        except Exception as e:
            log.error(f"Error reading workbook {file_id}: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {str(e)}",
            )

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error getting Excel metadata: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )
