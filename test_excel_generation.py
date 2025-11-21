#!/usr/bin/env python3
"""
Test script to generate a sample Excel file for testing the Excel artifact viewer.
This demonstrates what the code interpreter would do.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def create_sample_excel():
    """Create a sample Excel file with multiple sheets"""

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Report
    ws1 = wb.active
    ws1.title = "Sales Report"

    # Headers
    headers = ["Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Total"]
    ws1.append(headers)

    # Style headers
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    products = [
        ["Laptop", 15000, 18000, 22000, 25000],
        ["Mouse", 2500, 3000, 2800, 3500],
        ["Keyboard", 3500, 4000, 3800, 4200],
        ["Monitor", 12000, 13500, 15000, 16500],
        ["Headset", 4500, 5000, 5500, 6000],
    ]

    for product_data in products:
        row_data = product_data + [sum(product_data[1:])]
        ws1.append(row_data)

    # Add totals row
    ws1.append(["TOTAL", "=SUM(B2:B6)", "=SUM(C2:C6)", "=SUM(D2:D6)", "=SUM(E2:E6)", "=SUM(F2:F6)"])

    # Style totals row
    total_row = ws1.max_row
    for cell in ws1[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Adjust column widths
    ws1.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col].width = 12

    # Sheet 2: Customer Data
    ws2 = wb.create_sheet("Customer Data")

    customer_headers = ["Customer ID", "Name", "Email", "Phone", "Status"]
    ws2.append(customer_headers)

    # Style headers
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Sample customer data
    customers = [
        ["C001", "John Doe", "john@example.com", "555-0101", "Active"],
        ["C002", "Jane Smith", "jane@example.com", "555-0102", "Active"],
        ["C003", "Bob Johnson", "bob@example.com", "555-0103", "Inactive"],
        ["C004", "Alice Williams", "alice@example.com", "555-0104", "Active"],
        ["C005", "Charlie Brown", "charlie@example.com", "555-0105", "Active"],
    ]

    for customer in customers:
        ws2.append(customer)

    # Adjust column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 10

    # Sheet 3: Monthly Summary
    ws3 = wb.create_sheet("Monthly Summary")

    ws3.append(["Month", "Revenue", "Expenses", "Profit"])

    # Style headers
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Monthly data
    months = [
        ["January", 45000, 30000, 15000],
        ["February", 48000, 32000, 16000],
        ["March", 52000, 35000, 17000],
        ["April", 55000, 36000, 19000],
        ["May", 58000, 38000, 20000],
        ["June", 62000, 40000, 22000],
    ]

    for month_data in months:
        ws3.append(month_data)

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 15

    # Save the file
    output_path = "/tmp/sample_report.xlsx"
    wb.save(output_path)
    print(f"✅ Sample Excel file created: {output_path}")
    print(f"📊 File size: {os.path.getsize(output_path)} bytes")
    print(f"📄 Sheets: {', '.join(wb.sheetnames)}")

    return output_path

if __name__ == "__main__":
    create_sample_excel()
